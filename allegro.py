import tkinter as tk
from tkinter import simpledialog, messagebox
import webbrowser
import base64
import hashlib
import secrets
import string
import requests
import json
import os
import openpyxl
from tkinter import ttk
import datetime
import pandas as pd
import threading
import time
import random

# Wyczyszczenie konsoli
os.system("cls")

# Definicja klasy AllegroApp
class AllegroApp:
    def __init__(self, access_token=None):
        # Dodanie atrybutu do przechowywania wprowadzanej wiadomości
        self.message_entry = None

        # Dodatkowe pole do przechowywania ID ostatniej wiadomości
        self.last_message_id = None

        # Dodatkowe pole do przechowywania czasu ostatniego sprawdzenia wiadomości
        self.last_check_time = None

        # Pole przechowujące informacje o ostatniej wiadomości
        self.last_message_info = None

        # Zmienna do sprawdzenia, czy wątek powinien być uruchomiony
        self.should_run_message_check_thread = False


        self.wiadomosc_tekst_default = """Dziękujemy za wiadomość! Twój kontakt został odebrany, a my postaramy się odpowiedzieć najszybciej, jak to możliwe. Prosimy o cierpliwość.

Pozdrawiamy,
Zespół Obsługi Klienta"""
        # Uruchom wątek sprawdzający co 20 sekund
        self.start_message_check_thread()

        # Przypisanie stałych do zmiennych instancji
        self.CLIENT_ID = ""
        self.CLIENT_SECRET = ""
        self.REDIRECT_URI = "https://example.com"
        self.AUTH_URL = "https://allegro.pl/auth/oauth/authorize"
        self.TOKEN_URL = "https://allegro.pl/auth/oauth/token"
        self.API_URL = "https://api.allegro.pl/me"
        self.access_token = access_token

        # Inicjalizacja głównego okna tkinter
        self.root = tk.Tk()
        self.root.title("Aplikacja Allegro")

        # Obliczenia pozycji i rozmiaru okna
        window_width = 300
        window_height = 380
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")

        # Przycisk "Weryfikacja użytkownika" z przypisaną funkcją on_verify_user
        verify_button = ttk.Button(self.root, text="Logowanie", command=self.on_verify_user)
        verify_button.pack(pady=100)  # Zmniejsz wartość pady

        # Przycisk "Usuń zapisane konto" z przypisaną funkcją on_verify_user
        verify_button = ttk.Label(self.root, text="Czy chcesz usunąć zapisane konto?")
        verify_button.pack(pady=10)  # Zmniejsz wartość pady

        # Przycisk "Usuń zapisane konto" z przypisaną funkcją on_verify_user
        verify_button = ttk.Button(self.root, text="Usuń", command=self.usun_konto)
        verify_button.pack(pady=10)  # Zmniejsz wartość pady

        # Stopka
        label_opcje = ttk.Label(self.root, text="© 2023 Dakro Bosch Service Autor: Mike Boro", font=("Arial", 8))
        label_opcje.pack(pady=20, padx=20)

        # Ustawienie stylu przycisku
        style = ttk.Style()
        style.configure("TButton", font=("Helvetica", 9))  # Zmiana czcionki
        style.configure("TButton", padding=1)  # Dodanie wypełnienia
        style.configure("TButton", foreground="black")  # Zmiana koloru tekstu
        style.configure("TButton", background="#e1e1e1")  # Zmiana koloru tła
    
    def respond_to_new_message(self, interlocutor_login):
        try:
            # Adres URL do wysłania nowej wiadomości
            messages_url = "https://api.allegro.pl/messaging/messages"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
                'Content-Type': 'application/vnd.allegro.public.v1+json',
            }

            # Dane do wysłania
            data = {
                'recipient': {
                    'login': interlocutor_login,
                },
                'text': f"{self.wiadomosc_tekst_default}",
                'attachments': []
            }

            # Wyślij wiadomość
            response = requests.post(messages_url, headers=headers, json=data)
            response.raise_for_status()

            # Pobierz identyfikator wątku z odpowiedzi
            thread_id = response.json().get('thread', {}).get('id')

            # Zaktualizuj status wątku na odczytany
            self.mark_thread_as_read(thread_id)

            print(f"Wiadomość została wysłana do {interlocutor_login}!")

        except requests.exceptions.HTTPError as err:
            print(f"Błąd przy wysyłaniu wiadomości: {err}")

    def mark_thread_as_read(self, thread_id):
        try:
            # Adres URL do aktualizacji statusu wątku na odczytany
            mark_read_url = f"https://api.allegro.pl/messaging/threads/{thread_id}/read"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
                'Content-Type': 'application/vnd.allegro.public.v1+json',
            }

            # Dane do aktualizacji
            data = {
                'read': True,
            }

            # Wyślij zapytanie PUT
            response = requests.put(mark_read_url, headers=headers, json=data)
            response.raise_for_status()

            print(f"Status wątku {thread_id} został zaktualizowany na odczytany.")

        except requests.exceptions.HTTPError as err:
            print(f"Błąd przy aktualizacji statusu wątku: {err}")

    def start_message_check_thread(self):
        # Utwórz wątek do sprawdzania wiadomości
        self.message_check_thread = threading.Thread(target=self.check_for_messages, daemon=True)

        # Uruchom wątek tylko, jeśli self.should_run_message_check_thread to True
        if self.should_run_message_check_thread:
            self.message_check_thread.start()
            # Zaktualizuj label po uruchomieniu wątku
            self.automatyczne_wiadomosci.config(text="Automatyczne wiadomości włączone", fg="green")
    def on_button_press(self):
        # Ustaw flagę, że wątek powinien być uruchomiony
        self.should_run_message_check_thread = True

        # Uruchom funkcję, która rozpocznie wątek
        self.start_message_check_thread()
    def check_for_messages(self):
        while True:
            # Symulacja sprawdzania wiadomości losowo pomiędzy 60 a 130 sekund
            sleep_time = random.randint(60, 130)
            time.sleep(sleep_time)
            self.check_messages()
            print("Sprawdzono wiadomości!")

    def check_messages(self):
        try:
            # Adres URL do pobrania listy wątków użytkownika
            threads_url = "https://api.allegro.pl/messaging/threads"

            # Parametry zapytania
            params = {
                'limit': 10,  # Możesz dostosować ilość wątków do wyświetlenia
                'offset': 0,
            }

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
            }

            # Wykonaj zapytanie do API Allegro
            response = requests.get(threads_url, headers=headers, params=params)
            response.raise_for_status()

            # Pobierz dane wątków z odpowiedzi
            threads_data = response.json().get('threads', [])

            # Sprawdź, czy pojawiła się nowa wiadomość w istniejących wątkach
            for thread in threads_data:
                is_thread_read = thread.get('read')
                interlocutor_login = thread.get('interlocutor', {}).get('login')

                # Sprawdź, czy wątek jest nieprzeczytany
                if not is_thread_read:
                    print(f"Nieprzeczytany wątek od {interlocutor_login}!")
                    self.respond_to_new_message(interlocutor_login)
        except requests.exceptions.HTTPError as err:
            print(f"Błąd przy sprawdzaniu wiadomości: {err}")

    def usun_konto(self):
        # Wyświetlenie zapytania
        response = messagebox.askyesno("Usuwanie konta", "Czy na pewno chcesz usunąć zapisane konto?")

        if response:
            try:
                # Usunięcie pliku config.txt
                os.remove("config.txt")
                messagebox.showinfo("Usuwanie konta", "Konto zostało usunięte.")
                self.root.destroy()
                # Warunek sprawdzający, czy skrypt jest uruchamiany jako główny program
                if __name__ == "__main__":
                    try:
                        with open('config.txt', 'r') as file:
                            token = file.read().strip()
                    except:
                        with open('config.txt', 'w') as file:
                            pass
                    with open('config.txt', 'r') as file:
                            token = file.read().strip()
                    app = AllegroApp(token)
                    app.run()
            except FileNotFoundError:
                messagebox.showinfo("Usuwanie konta", "Nie znaleziono pliku config.txt.")
        else:
            messagebox.showinfo("Usuwanie konta", "Operacja usunięcia konta została anulowana.")

    def show_thread_messages(self, thread_id):
        try:
            # Adres URL do pobrania listy wiadomości dla danego wątku
            messages_url = f"https://api.allegro.pl/messaging/threads/{thread_id}/messages"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
            }

            # Wykonaj zapytanie do API Allegro
            response = requests.get(messages_url, headers=headers)
            response.raise_for_status()

            # Pobierz dane wiadomości z odpowiedzi
            messages_data = response.json().get('messages', [])
            # Odwróć listę, aby mieć najstarsze wiadomości na początku
            messages_data.reverse()

            # Utwórz nowe okno dla treści wątku
            thread_window = tk.Toplevel(self.root)
            thread_window.title("Treść Wątku")
            self.thread_window = thread_window

            # Utwórz pole tekstowe do wyświetlenia treści wątku
            text_widget = tk.Text(thread_window, wrap="word", width=60, height=20)
            text_widget.pack(padx=10, pady=10)

            # Dodaj pole tekstowe na dole okna
            self.message_entry = tk.Text(thread_window, width=60, height=5)
            self.message_entry.pack(pady=5, padx=10)

            # Dodaj przycisk "Wyślij" obok pola tekstowego
            send_button = ttk.Button(thread_window, text="Wyślij", command=self.send_message, width=10)
            send_button.pack(pady=15, padx=10)

            # Wypełnij pole tekstowe danymi
            for message in messages_data:
                author_login = message.get('author', {}).get('login')
                message_text = message.get('text')
                created_at = message.get('createdAt')

                # Dodaj informacje o wiadomości do pola tekstowego
                timestamp = f"{created_at.split('T')[1][:5]} | {created_at.split('T')[0]}"
                text_widget.insert(tk.END, f"{author_login} ({timestamp}):\n{message_text}\n\n")

            # Przewiń pole tekstowe na sam dół
            text_widget.see(tk.END)

            text_widget.config(state=tk.DISABLED)  # Zablokuj edycję tekstu

        except requests.exceptions.HTTPError as err:
            messagebox.showinfo("Błąd", f"Błąd przy pobieraniu treści wątku: {err}")

    def send_message(self):
        try:
            # Pobierz wprowadzoną wiadomość
            message_text = self.message_entry.get("1.0", tk.END)

            # Sprawdź, czy wybrano wątek
            selected_thread_id = self.get_selected_thread_id()
            if selected_thread_id is None:
                messagebox.showinfo("Błąd", "Wybierz wątek, aby wysłać wiadomość.")
                return
            
            # Adres URL do wysłania wiadomości w wybranym wątku
            messages_url = f"https://api.allegro.pl/messaging/threads/{selected_thread_id}/messages"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
                'Content-Type': 'application/vnd.allegro.public.v1+json',
            }

            # Dane do wysłania
            data = {
                'text': message_text,
                'attachments': []
            }

            # Wyślij wiadomość
            response = requests.post(messages_url, headers=headers, json=data)
            response.raise_for_status()

            # Wyczyść pole tekstowe
            self.message_entry.delete("1.0", tk.END)

            messagebox.showinfo("Sukces", "Wiadomość została wysłana pomyślnie!")
            
            self.thread_window.destroy()
        except requests.exceptions.HTTPError as err:
            messagebox.showinfo("Błąd", f"Błąd przy wysyłaniu wiadomości: {err}")

    def get_selected_thread_id(self):
        selected_item = self.tree.focus()
        if selected_item:
            thread_id = self.tree.item(selected_item)['values'][1]
            return thread_id
        else:
            return None
        
    def wiadomosc_edytuj(self):
        # Utwórz nowe okno do wprowadzenia treści wiadomości
        wiadomosc_edit = tk.Toplevel(self.root)
        self.wiadomosc_edit = wiadomosc_edit
        wiadomosc_edit.title("Edytuj automatyczną wiadomość")
        wiadomosc_edit.geometry("700x350")

        # Utwórz pole tekstowe do wprowadzenia treści wiadomości
        self.wiadomosc_tekst = tk.Text(wiadomosc_edit, wrap="word", width=80, height=15)
        self.wiadomosc_tekst.insert(tk.END, self.wiadomosc_tekst_default)
        self.wiadomosc_tekst.pack(padx=10, pady=10)

        # Dodaj przycisk "Zapisz" obok pola tekstowego
        send_button_all = ttk.Button(wiadomosc_edit, text="Zapisz", command=self.wiadomosc_edytuj_zapisz)
        send_button_all.pack(pady=10)

    def wiadomosc_edytuj_zapisz(self):
        # Pobierz wprowadzoną wiadomość
        message_text_all = self.wiadomosc_tekst.get("1.0", tk.END).strip()

        # Sprawdź, czy wprowadzono treść wiadomości
        if not message_text_all:
            messagebox.showinfo("Błąd", "Wprowadź treść wiadomości.")
            return

        # Tutaj możesz zapisywać wartość do self.wiadomosc_tekst, jeśli to konieczne
        self.wiadomosc_tekst_default = message_text_all
        print(f"Wprowadzona treść wiadomości: {message_text_all}")
        messagebox.showinfo("Zapisano", "Pomyślnie zapisano treść!")
        self.wiadomosc_edit.destroy()

    def wiad_fun(self):
        try:
            # Adres URL do pobrania listy wątków użytkownika
            threads_url = "https://api.allegro.pl/messaging/threads"

            # Parametry zapytania
            params = {
                'limit': 20,  # Możesz dostosować ilość wątków do wyświetlenia
                'offset': 0,
            }

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
            }

            # Wykonaj zapytanie do API Allegro
            response = requests.get(threads_url, headers=headers, params=params)
            response.raise_for_status()

            # Pobierz dane wątków z odpowiedzi
            threads_data = response.json().get('threads', [])
            
            # Utwórz nowe okno dla tabeli
            threads_window = tk.Toplevel(self.root)
            threads_window.title("Centrum Wiadomości")

            # Label
            open_message_button_label_b = ttk.Label(threads_window, text="Wyślij wiadomość do wszystkich:")
            open_message_button_label_b.pack(padx=5,pady=5)

            # Przycisk "Wyślij do wszystkich" z przypisaną funkcją on_send_to_all
            send_to_all_button = ttk.Button(threads_window, text="Wyślij", command=self.on_send_to_all)
            send_to_all_button.pack(padx=5, pady=5)

            # Label
            open_message_button_label_a = ttk.Label(threads_window, text="Uruchom automatyczne wiadomości:")
            open_message_button_label_a.pack(padx=5,pady=5)

            # Przycisk "Automatyczne Wiadomości"
            send_to_all_button = ttk.Button(threads_window, text="Uruchom", command=self.on_button_press)
            send_to_all_button.pack(padx=5, pady=5)

            # Label
            open_message_button_label = ttk.Label(threads_window, text="Edytuj treść automatycznej wiadomości:")
            open_message_button_label.pack(padx=5,pady=5)

            # Przycisk do otwierania okna wprowadzania wiadomości
            open_message_button = ttk.Button(threads_window, text="Edytuj", command=self.wiadomosc_edytuj)
            open_message_button.pack(padx=5, pady=5)
            
            # Utwórz tabelę Treeview
            tree = ttk.Treeview(threads_window)
            tree["columns"] = ("Nadawca", "Wątek", "Czas")
            tree.column("#0", width=0, stretch=tk.NO)
            tree.column("Nadawca", anchor=tk.W, width=150)
            tree.column("Wątek", anchor=tk.W, width=400)
            tree.column("Czas", anchor=tk.W, width=150)

            tree.heading("#0", text="", anchor=tk.W)
            tree.heading("Nadawca", text="Nadawca", anchor=tk.W)
            tree.heading("Wątek", text="Wątek", anchor=tk.W)
            tree.heading("Czas", text="Data i czas", anchor=tk.W)
            self.tree = tree

            # Utwórz pionowy suwak
            scrollbar = ttk.Scrollbar(threads_window, orient='vertical', command=tree.yview)
            scrollbar.pack(side='right', fill='y')

            # Połącz suwak ze wskazanym Treeview
            tree.configure(yscrollcommand=scrollbar.set)

            # Wypełnij tabelę danymi
            for thread in threads_data:
                thread_id = thread.get('id')
                interlocutor_login = thread.get('interlocutor', {}).get('login')
                last_message_datetime = thread.get('lastMessageDateTime')

                # Przekształć string daty do obiektu datetime
                timestamp = datetime.datetime.strptime(last_message_datetime, '%Y-%m-%dT%H:%M:%S.%fZ')

                # Sformatuj datę do oczekiwanego formatu
                formatted_timestamp = timestamp.strftime('%d.%m.%Y %H:%M:%S')

                # Dodaj wiersz do tabeli
                tree.insert("", "end", values=(interlocutor_login, thread_id, formatted_timestamp))

            # Podwójne
            tree.bind("<Double-1>", lambda event: self.show_thread_messages(tree.item(tree.focus())['values'][1]))
            tree.pack(expand=True, fill="both")

        except requests.exceptions.HTTPError as err:
            messagebox.showinfo("Błąd", f"Błąd przy pobieraniu listy wątków: {err}")

    def on_send_to_all(self):
        # Utwórz nowe okno do wprowadzenia treści wiadomości
        send_to_all_window = tk.Toplevel(self.root)
        send_to_all_window.title("Wyślij do wszystkich")
        self.send_to_all_window = send_to_all_window

        # Utwórz pole tekstowe do wprowadzenia treści wiadomości
        message_entry_all = tk.Text(send_to_all_window, wrap="word", width=40, height=10)
        message_entry_all.pack(padx=10, pady=10)

        # Dodaj przycisk "Wyślij" obok pola tekstowego
        send_button_all = ttk.Button(send_to_all_window, text="Wyślij", command=lambda: self.send_to_all(message_entry_all))
        send_button_all.pack(pady=10)

    def send_to_all(self, message_entry_all):
        try:
            # Pobierz wprowadzoną wiadomość
            message_text_all = message_entry_all.get("1.0", tk.END).strip()

            # Sprawdź, czy wprowadzono treść wiadomości
            if not message_text_all:
                messagebox.showinfo("Błąd", "Wprowadź treść wiadomości.")
                return

            # Adres URL do wysłania wiadomości do wszystkich wątków
            threads_url = "https://api.allegro.pl/messaging/threads"

            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
                'Content-Type': 'application/vnd.allegro.public.v1+json',
            }

            # Pętla do pobrania listy wątków
            response = requests.get(threads_url, headers=headers, params={'limit': 20, 'offset': 0})
            response.raise_for_status()
            threads_data = response.json().get('threads', [])

            # Pętla do wysyłania wiadomości do każdego wątku
            for thread in threads_data:
                thread_id = thread.get('id')

                # Adres URL do wysłania wiadomości w danym wątku
                messages_url = f"https://api.allegro.pl/messaging/threads/{thread_id}/messages"

                # Dane do wysłania
                data = {
                    'text': message_text_all,
                    'attachments': []
                }

                # Wyślij wiadomość
                response = requests.post(messages_url, headers=headers, json=data)
                response.raise_for_status()

            # Zamknij okno po wysłaniu wiadomości do wszystkich wątków
            messagebox.showinfo("Sukces", "Wiadomość została wysłana do wszystkich wątków!")
            message_entry_all.delete("1.0", tk.END)  # Wyczyść pole tekstowe
            self.send_to_all_window.destroy()

        except requests.exceptions.HTTPError as err:
            messagebox.showinfo("Błąd", f"Błąd przy wysyłaniu wiadomości do wszystkich wątków: {err}")

    def save_all_offers_to_excel(self, active_offers):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Aktywne Oferty"
            
            # Nagłówki kolumn
            ws['A1'] = 'Tytuł Oferty'
            ws['B1'] = 'Liczba Sztuk'

            # Wypełnij danymi
            for index, offer in enumerate(active_offers, start=2):
                ws.cell(row=index, column=1, value=offer['name'])
                ws.cell(row=index, column=2, value=offer['stock']['available'])

            # Zapisz do pliku
            wb.save('aktywne_oferty.xlsx')
            messagebox.showinfo("Zapisano", "Aktywne oferty zostały zapisane do pliku!")

        except Exception as e:
            messagebox.showinfo("Błąd", f"Błąd przy zapisywaniu aktywnych ofert do pliku: {e}")

    # Generowanie losowego code_verifier
    def generate_code_verifier(self):
        return ''.join((secrets.choice(string.ascii_letters) for i in range(40)))

    # Generowanie code_challenge na podstawie code_verifier
    def generate_code_challenge(self, code_verifier):
        hashed = hashlib.sha256(code_verifier.encode('utf-8')).digest()
        base64_encoded = base64.urlsafe_b64encode(hashed).decode('utf-8')
        return base64_encoded.replace('=', '')

    # Otwieranie strony logowania Allegro i zwracanie code_verifier
    def open_allegro_login(self):
        code_verifier = self.generate_code_verifier()
        code_challenge = self.generate_code_challenge(code_verifier)
        auth_url = f"{self.AUTH_URL}?response_type=code&client_id={self.CLIENT_ID}&redirect_uri={self.REDIRECT_URI}" \
                   f"&code_challenge_method=S256&code_challenge={code_challenge}"
        webbrowser.open(auth_url)
        return code_verifier

    # Pobieranie informacji o koncie z Allegro
    def get_account_info(self):
        try:
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
            }
            response = requests.get(self.API_URL, headers=headers)
            response.raise_for_status()  # Raises HTTPError for bad responses
            return response.json()
        except requests.exceptions.HTTPError as err:
            if err.response.status_code == 401:  # Unauthorized (invalid/expired token)
                messagebox.showinfo("Błąd autoryzacji", "Token dostępu wygasł lub jest nieprawidłowy!")
                # Start the verification process again
                code_verifier = self.open_allegro_login()
                authorization_code = simpledialog.askstring("Wprowadź kod autoryzacyjny",
                                                            "Podaj kod autoryzacyjny:")
                self.get_access_token(authorization_code, code_verifier)
                return self.get_account_info()  # Retry after reauthorization
            else:
                messagebox.showinfo("Błąd", f"Błąd przy pobieraniu informacji o koncie: {err}")
                return {}

    # Pobieranie access token na podstawie authorization code
    def get_access_token(self, authorization_code, code_verifier):
        try:
            data = {'grant_type': 'authorization_code', 'code': authorization_code,
                    'redirect_uri': self.REDIRECT_URI, 'code_verifier': code_verifier}
            access_token_response = requests.post(self.TOKEN_URL, data=data, verify=False, allow_redirects=False)
            response_body = json.loads(access_token_response.text)
            self.access_token = response_body['access_token']
            
            #print(self.access_token)
            with open('config.txt', 'w') as file:
                file.write(self.access_token)

        except requests.exceptions.HTTPError as err:
            raise SystemExit(err)

    # Obsługa weryfikacji użytkownika
    def on_verify_user(self):
        if self.access_token is None:
            # Otwórz stronę logowania Allegro i pobierz code_verifier
            code_verifier = self.open_allegro_login()

            # Pobierz authorization code od użytkownika
            authorization_code = simpledialog.askstring(" ",
                                                        "Podaj kod autoryzacyjny:")
            # Sprawdź, czy użytkownik anulował
            if authorization_code is None:
                messagebox.showinfo("Logowanie", f"Anulowano logowanie!")
                return  # Przerwij proces weryfikacji

            # Pobierz access token
            self.get_access_token(authorization_code, code_verifier)

        # Pobierz informacje o koncie
        account_info = self.get_account_info()

        # Wyświetlanie informacji o koncie w tym samym oknie
        self.show_account_info(account_info)

    # Wyświetlanie informacji o koncie w oknie aplikacji
    def show_account_info(self, account_info):
        # Czyszczenie poprzednich widgetów z okna głównego
        for widget in self.root.winfo_children():
            widget.destroy()

        # Ramka na etykiety
        labels_frame = tk.Frame(self.root, padx=20, pady=20)
        labels_frame.pack()

        # Wyświetlanie wybranych informacji o koncie
        info_keys = ['id', 'login', 'firstName', 'lastName', 'email']
        for key in info_keys:
            translated_key = self.translate_polish_words(key)
            value = account_info[key]
            cleaned_value = "Brak" if value is None else str(value).capitalize()
            label_text = f"{translated_key}: {cleaned_value}"
            label = tk.Label(labels_frame, text=label_text, anchor="w", justify="left", wraplength=400)
            label.pack(side="top", fill="x")

        # Label informacyjny
        self.automatyczne_wiadomosci = tk.Label(self.root, text="Automatyczne wiadomości wyłączone", anchor="w",
                                               justify="left", wraplength=400, padx=5, pady=5, fg="red")
        self.automatyczne_wiadomosci.pack()

        # Przycisk "Saldo Konta"
        saldo_button = ttk.Button(self.root, text="Saldo konta", command=self.show_balance, width=20)
        saldo_button.pack(padx=5, pady=10)

        # Przycisk Oferty
        offers_button = ttk.Button(self.root, text="Pobierz aktywne oferty", command=self.show_all_offers, width=20)
        offers_button.pack(padx=5, pady=10)

        # Przycisk Pokaż wiadomości
        wiad_button = ttk.Button(self.root, text="Wyświetl wiadomości", command=self.wiad_fun, width=20)
        wiad_button.pack(padx=5, pady=10)
        
        # Przycisk Pokaż wiadomości
        wiad_button = ttk.Button(self.root, text="Sprawdź cenę", command=self.price_check, width=20)
        wiad_button.pack(padx=5, pady=10)

        # Stopka
        label_opcje = ttk.Label(self.root, text="© 2023 Dakro Bosch Service Autor: Mike Boro", font=("Arial", 8))
        label_opcje.pack(pady=2, padx=2)

    def price_check(self):
        phrases_data = simpledialog.askstring(" ", "Podaj frazy oddzielone przecinkiem:")
        # Sprawdź, czy użytkownik kliknął Anuluj
        if phrases_data is None:
            return

        phrases = [phrase.strip() for phrase in phrases_data.split(',')]
        results_window = tk.Toplevel()
        results_window.title("Wyniki Wyszukiwania")
        results_window.geometry("600x400")

        results_text = tk.Text(results_window, wrap="word", width=80, height=20)
        results_text.pack(padx=10, pady=10)

        export_button = ttk.Button(results_window, text="Eksportuj wszystko", command=lambda: self.export_all_to_excel(phrases, results_text), width=20)
        export_button.pack(padx=10, pady=10)

        for phrase_data in phrases:
            try:
                # Adres URL do pobrania listy wyszukanych aukcji
                offers_url = "https://api.allegro.pl/offers/listing"

                # Parametry zapytania
                params = {
                    'limit': 60,
                    'offset': 0,
                    'phrase': f"{phrase_data}",
                    'searchMode': "DESCRIPTIONS",
                    'parameter.11323': 11323_2,
                }

                headers = {
                    'Authorization': f'Bearer {self.access_token}',
                    'Accept': 'application/vnd.allegro.public.v1+json',
                }

                # Wykonaj zapytanie do API Allegro
                response = requests.get(offers_url, headers=headers, params=params)
                response.raise_for_status()

                # Pobierz dane ofert z odpowiedzi
                offers_data = [
                    {'id': '123456789', 'sellingMode': {'price': {'amount': 99.99}}, 'name': 'Przykładowa oferta 1'},
                    {'id': '987654321', 'sellingMode': {'price': {'amount': 150.0}}, 'name': 'Przykładowa oferta 2'},
                    {'id': '567890123', 'sellingMode': {'price': {'amount': 322.20}}, 'name': 'Przykładowa oferta 3'},
                    {'id': '456789012', 'sellingMode': {'price': {'amount': 80.0}}, 'name': 'Przykładowa oferta 4'},
                    {'id': '234567890', 'sellingMode': {'price': {'amount': 120.0}}, 'name': 'Przykładowa oferta 5'},
                ]
                offers_data = response.json().get('items', [])
                # Utwórz etykietę z informacją o frazie
                results_text.insert(tk.END, f"\nWyszukiwana fraza: {phrase_data}\n")

                total_price = 0
                min_price = float('inf')
                max_price = float('-inf')
                len_offers = len(offers_data)

                for offer in offers_data:
                    offer_price = offer.get('sellingMode', {}).get('price', {}).get('amount')

                    # Aktualizuj wartości dla obliczeń
                    total_price += offer_price
                    min_price = min(min_price, offer_price)
                    max_price = max(max_price, offer_price)

                # Oblicz średnią cenę
                average_price = total_price / len_offers if len_offers else 0

                # Dodaj etykiety z informacjami do pola tekstowego
                results_text.insert(tk.END, f"Ilość aukcji: {len_offers}\n")
                results_text.insert(tk.END, f"Najwyższa cena: {max_price} PLN\nNajniższa cena: {min_price} PLN\n"
                                            f"Średnia cena: {average_price:.2f} PLN\n\n")

            except requests.exceptions.HTTPError as err:
                messagebox.showinfo("Błąd", f"Błąd przy pobieraniu listy ofert: {err}")

    def export_all_to_excel(self, phrases, results_text):
        all_data = {
            "Fraza": [],
            "Najwyższa Cena": [],
            "Najniższa Cena": [],
            "Średnia Cena": [],
            "Ilość Aukcji": [],
        }

        for phrase_data in phrases:
            try:
                # Adres URL do pobrania listy wyszukanych aukcji
                offers_url = "https://api.allegro.pl/offers/listing"

                # Parametry zapytania
                params = {
                    'limit': 60,
                    'offset': 0,
                    'phrase': f"{phrase_data}",
                    'searchMode': "DESCRIPTIONS",
                }

                headers = {
                    'Authorization': f'Bearer {self.access_token}',
                    'Accept': 'application/vnd.allegro.public.v1+json',
                }

                # Wykonaj zapytanie do API Allegro
                response = requests.get(offers_url, headers=headers, params=params)
                response.raise_for_status()

                # Pobierz dane ofert z odpowiedzi
                offers_data = [
                    {'id': '123456789', 'sellingMode': {'price': {'amount': 99.99}}, 'name': 'Przykładowa oferta 1'},
                    {'id': '987654321', 'sellingMode': {'price': {'amount': 150.0}}, 'name': 'Przykładowa oferta 2'},
                    {'id': '567890123', 'sellingMode': {'price': {'amount': 322.20}}, 'name': 'Przykładowa oferta 3'},
                    {'id': '456789012', 'sellingMode': {'price': {'amount': 80.0}}, 'name': 'Przykładowa oferta 4'},
                    {'id': '234567890', 'sellingMode': {'price': {'amount': 120.0}}, 'name': 'Przykładowa oferta 5'},
                ]
                offers_data = response.json().get('items', [])
                
                total_price = 0
                min_price = float('inf')
                max_price = float('-inf')
                len_offers = len(offers_data)

                for offer in offers_data:
                    offer_price = offer.get('sellingMode', {}).get('price', {}).get('amount')

                    # Aktualizuj wartości dla obliczeń
                    total_price += offer_price
                    min_price = min(min_price, offer_price)
                    max_price = max(max_price, offer_price)

                # Oblicz średnią cenę
                average_price = total_price / len_offers if len_offers else 0

                # Dodaj dane do słownika
                all_data["Fraza"].append(phrase_data)
                all_data["Najwyższa Cena"].append(max_price)
                all_data["Najniższa Cena"].append(min_price)
                all_data["Średnia Cena"].append(average_price)
                all_data["Ilość Aukcji"].append(len_offers)

            except requests.exceptions.HTTPError as err:
                messagebox.showinfo("Błąd", f"Błąd przy pobieraniu listy ofert: {err}")

        # Eksportuj dane do jednego pliku Excel
        df = pd.DataFrame(all_data)
        df.to_excel(f"Analiza Cen.xlsx", index=False)
        messagebox.showinfo("Eksport zakończony", f"Wszystkie wyniki zostały wyeksportowane do pliku Excel - Analaiza Cen.xlsx")

    def show_all_offers(self):
        try:
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Accept': 'application/vnd.allegro.public.v1+json',
            }

            # Parametry zapytania
            params = {
                'limit': 1000,  # Maksymalna liczba ofert na stronie
            }

            # Zainicjowanie listy, aby przechowywać aktywne oferty z ilością > 0
            active_offers = []

            # Pętla do obsługi paginacji
            while True:
                response = requests.get("https://api.allegro.pl/sale/offers", headers=headers, params=params)
                response.raise_for_status()  # Raises HTTPError for bad responses
                offers_data = response.json()

                # Sprawdź, czy są oferty w odpowiedzi
                if 'offers' in offers_data:
                    # Dodaj aktywne oferty z ilością > 0 do listy
                    for offer in offers_data['offers']:
                        if offer.get('publication') and offer['publication'].get('status') == 'ACTIVE' and offer.get('stock') and offer['stock'].get('available') > 0:
                            active_offers.append(offer)

                # Sprawdź, czy są więcej strony z ofertami
                if '_links' in offers_data and 'next' in offers_data['_links']:
                    params['offset'] = int(offers_data['_links']['next']['href'].split('offset=')[1])
                else:
                    break

            # Jeżeli lista aktywnych ofert jest pusta, to nie znaleziono pasujących ofert
            if not active_offers:
                messagebox.showinfo("Nie znaleziono", "Brak aktywnych aukcji z ilością większą niż 0!")
            else:
                self.save_all_offers_to_excel(active_offers)

        except requests.exceptions.HTTPError as err:
            messagebox.showinfo("Błąd", f"Błąd przy pobieraniu listy aktywnych ofert: {err}")

    # Wyświetlanie salda konta w oknie aplikacji
    def show_balance(self):
        balance = self.get_account_balance()
        if balance is None:
            messagebox.showinfo("Saldo konta", "Saldo konta wynosi: 0")
        else:
            messagebox.showinfo("Saldo konta", f"Saldo konta wynosi: {balance}")

    # Pobieranie salda konta z Allegro
    def get_account_balance(self):
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'Accept': 'application/vnd.allegro.public.v1+json',
        }
        response = requests.get(self.API_URL, headers=headers)
        return response.json().get('billing')

    # Tłumaczenie nazw polskich słów kluczowych na angielskie
    def translate_polish_words(self, word):
        translations = {
            'id': 'ID',
            'login': 'Nazwa użytkownika',
            'firstName': 'Imię',
            'lastName': 'Nazwisko',
            'email': 'E-Mail',
            'baseMarketplace': 'Baza rynku',
            'company': 'Firma',
            'features': 'Funkcje',
        }
        return translations.get(word, word.capitalize())

    # Uruchomienie pętli głównej aplikacji
    def run(self):
        self.root.mainloop()
    
# Warunek sprawdzający, czy skrypt jest uruchamiany jako główny program
if __name__ == "__main__":
    try:
        with open('config.txt', 'r') as file:
            token = file.read().strip()
    except:
        with open('config.txt', 'w') as file:
            pass
    with open('config.txt', 'r') as file:
            token = file.read().strip()
    app = AllegroApp(token)
    app.run()
